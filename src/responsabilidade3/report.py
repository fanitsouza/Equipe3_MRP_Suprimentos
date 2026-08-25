from __future__ import annotations

import logging
import time
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .input_contract import ContractValidationError
from .models import MRPResult
from .severity import Severity
from .structured_logging import log_event


REPORT_HEADERS = [
    "Fornecedor",
    "Material",
    "Estoque",
    "Necessidade",
    "Capacidade",
    "Prazo_Dias",
    "Status_Validacao",
    "Observacao",
]


class ReportGenerationError(RuntimeError):
    pass


def generate_excel_report(
    results: Iterable[MRPResult],
    output_path: str | Path,
    template_path: str | Path | None = None,
    logger: logging.Logger | None = None,
    execution_id: str | None = None,
) -> Path:
    start = time.perf_counter()
    output = Path(output_path)
    template = Path(template_path) if template_path is not None else None

    try:
        safe_results = _validate_results(results)
        _ensure_output_does_not_overwrite_template(output, template)
        workbook = _load_workbook(template)
        worksheet = (
            workbook["Relatorio"]
            if "Relatorio" in workbook.sheetnames
            else workbook.active
        )
        worksheet.title = "Relatorio"

        _write_headers(worksheet)
        _clear_existing_data(worksheet)

        for row_index, result in enumerate(safe_results, start=2):
            row = result.to_report_row()
            for column_index, header in enumerate(REPORT_HEADERS, start=1):
                value = _sanitize_excel_value(row.get(header))
                worksheet.cell(row=row_index, column=column_index, value=value)

        output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output)
    except Exception as exc:
        duration_ms = int((time.perf_counter() - start) * 1000)
        if logger is not None:
            log_event(
                logger,
                level=Severity.ERROR,
                event="report_generation_failed",
                module="excel_report",
                message="Falha ao gerar relatorio Excel",
                execution_id=execution_id,
                error_type=type(exc).__name__,
                duration_ms=duration_ms,
            )
        if isinstance(exc, ReportGenerationError):
            raise
        raise ReportGenerationError(str(exc)) from exc

    duration_ms = int((time.perf_counter() - start) * 1000)
    if logger is not None:
        log_event(
            logger,
            level=Severity.INFO,
            event="report_generated",
            module="excel_report",
            message="Relatorio gerado com sucesso",
            execution_id=execution_id,
            duration_ms=duration_ms,
            row_count=len(safe_results),
            output_path=str(output),
        )
    return output


def _load_workbook(template_path: Path | None) -> Workbook:
    if template_path is None:
        return Workbook()

    if template_path.exists():
        return load_workbook(template_path)

    return Workbook()


def _write_headers(worksheet) -> None:
    for column_index, header in enumerate(REPORT_HEADERS, start=1):
        worksheet.cell(row=1, column=column_index, value=header)


def _clear_existing_data(worksheet) -> None:
    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)


def _validate_results(results: Iterable[MRPResult]) -> list[MRPResult]:
    validated: list[MRPResult] = []
    for index, result in enumerate(results, start=1):
        if not isinstance(result, MRPResult):
            raise ReportGenerationError(
                f"Entrada {index} nao segue o contrato MRPInputContract"
            )
        try:
            validated.append(MRPResult(**result.to_dict()))
        except ContractValidationError as exc:
            raise ReportGenerationError(f"Entrada {index} invalida: {exc}") from exc
    return validated


def _sanitize_excel_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return "'" + value
    return value


def _ensure_output_does_not_overwrite_template(
    output_path: Path,
    template_path: Path | None,
) -> None:
    if template_path is None:
        return
    if output_path.resolve() == template_path.resolve():
        raise ReportGenerationError("Arquivo de saida nao pode sobrescrever o template")
