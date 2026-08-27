import json
from pathlib import Path
from openpyxl import load_workbook
from src.config import Settings
from src.pipeline import executar_processo


def test_full_pipeline_integrated(tmp_path: Path):
    source_dir = Path("Source").resolve()
    output_dir = tmp_path / "output"
    alert_file = tmp_path / "logs" / "alerts.jsonl"

    settings = Settings(
        source_dir=source_dir,
        output_dir=output_dir,
        alert_file=alert_file,
        grp_url=(source_dir / "web" / "grp_fake.html").as_uri(),
        grp_user="aluno",
        grp_password="avaliacao2026",
        timezone="America/Manaus",
    )

    execution = executar_processo(settings)

    # 1. Validar execucao
    assert execution.execution_id is not None
    assert not execution.usou_fallback
    assert len(execution.resultados) == 4

    # 2. Validar relatorio Excel
    assert execution.relatorio.exists()
    wb = load_workbook(execution.relatorio, data_only=True)
    sheet = wb["Relatorio"]
    assert sheet.max_row == 5  # 1 cabecalho + 4 materiais (MAT001 a MAT004)

    # 3. Validar regra Fornecedor C
    mat003_row = [r for r in execution.resultados if r.material == "MAT003"][0]
    assert mat003_row.status_validacao == "AGUARDANDO_VALIDACAO_HUMANA"
    assert mat003_row.quantidade_comprar is None

    # 4. Validar cache de resiliencia
    cache_file = output_dir / "ultimo_mrp_valido.json"
    assert cache_file.exists()
    cached_data = json.loads(cache_file.read_text(encoding="utf-8"))
    assert len(cached_data) == 4

    # 5. Validar arquivo de alertas
    assert alert_file.exists()
    alert_lines = alert_file.read_text(encoding="utf-8").strip().splitlines()
    assert any("VALIDACAO_HUMANA_NECESSARIA" in line for line in alert_lines)
