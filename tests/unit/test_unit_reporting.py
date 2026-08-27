from pathlib import Path
from openpyxl import load_workbook
from src.models.data import StockRecord
from src.mrp.models import ResultadoMRP, StatusMRP, Severidade
from src.reporting.adapter import adapt_engine_results
from src.reporting.excel import generate_excel_report
from src.reporting.models import MRPInputContract


def test_adapt_engine_results():
    engine_results = [
        ResultadoMRP(
            material="MAT001",
            fornecedor="Fornecedor A",
            necessidade_calculada=100,
            quantidade_sugerida=100,
            capacidade_considerada=200,
            prazo_considerado_dias=5,
            status=StatusMRP.COMPRA_SUGERIDA,
            severidade=Severidade.INFO,
            requer_validacao_humana=False,
            mensagem="Compra normal",
        )
    ]
    stock_records = [
        StockRecord(material="MAT001", stock=100.0, weekly_demand=150.0, safety_stock=50.0)
    ]

    adapted = adapt_engine_results(engine_results, stock_records)
    assert len(adapted) == 1
    item = adapted[0]
    assert isinstance(item, MRPInputContract)
    assert item.material == "MAT001"
    assert item.estoque_atual == 100
    assert item.necessidade == 100
    assert item.quantidade_comprar == 100
    assert item.status_validacao == "COMPRA_SUGERIDA"


def test_generate_excel_report(tmp_path: Path):
    output_file = tmp_path / "relatorio_teste.xlsx"
    items = [
        MRPInputContract("MAT001", "Fornecedor A", 100, 150, 100, 200, 5, "COMPRA_SUGERIDA", "OK"),
        MRPInputContract("MAT003", "Fornecedor C", 50, 210, None, 100, 14, "AGUARDANDO_VALIDACAO_HUMANA", "Alfandega"),
    ]
    generate_excel_report(items, output_path=output_file)
    assert output_file.exists()

    wb = load_workbook(output_file, data_only=True)
    assert "Relatorio" in wb.sheetnames
    sheet = wb["Relatorio"]
    assert sheet.max_row == 3  # 1 cabecalho + 2 linhas de dados
    assert sheet.cell(1, 1).value == "Fornecedor"
    assert sheet.cell(1, 2).value == "Material"
    assert sheet.cell(2, 1).value == "Fornecedor A"
    assert sheet.cell(2, 2).value == "MAT001"
    assert sheet.cell(3, 1).value == "Fornecedor C"
    assert sheet.cell(3, 2).value == "MAT003"
