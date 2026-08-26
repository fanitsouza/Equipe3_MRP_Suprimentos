from datetime import date

from openpyxl import load_workbook

from src.config import Settings
from src.models.data import NFPRecord, StockRecord, SupplierRecord, SupplierUpdate
from src.pipeline import executar_processo
from src.resilience import CircuitBreaker


def _settings(tmp_path) -> Settings:
    return Settings(
        source_dir=tmp_path,
        output_dir=tmp_path / "output",
        alert_file=tmp_path / "logs" / "alerts.jsonl",
        grp_url="http://localhost/grp",
        grp_user="aluno",
        grp_password="senha",
        timezone="America/Manaus",
    )


def _coleta() -> dict:
    suppliers = [
        SupplierRecord("Fornecedor A", "MAT001", 200.0, 5, 100.0, "Ativo"),
        SupplierRecord("Fornecedor C", "MAT003", 150.0, 10, 110.0, "Ativo"),
    ]
    return {
        "stock": [
            StockRecord("MAT001", 100.0, 150.0, 50.0),
            StockRecord("MAT003", 50.0, 180.0, 80.0),
        ],
        "supplier_csv": suppliers,
        "grp_web": list(suppliers),
        "nfp": NFPRecord(
            "10452", "Fornecedor A", "MAT001", 180.0, date(2026, 3, 1), 100.0
        ),
        "supplier_update": SupplierUpdate(
            "Fornecedor C", "MAT003", 150.0, 100.0, 10, 14, "Alfândega"
        ),
    }


def test_pipeline_gera_relatorio_cache_e_alerta(monkeypatch, tmp_path) -> None:
    monkeypatch.setattr("src.pipeline.collect_all", lambda settings: _coleta())
    settings = _settings(tmp_path)

    execution = executar_processo(settings, CircuitBreaker(failure_threshold=3))

    assert execution.relatorio.exists()
    assert not execution.usou_fallback
    assert (settings.output_dir / "ultimo_mrp_valido.json").exists()
    assert "VALIDACAO_HUMANA_NECESSARIA" in settings.alert_file.read_text("utf-8")
    sheet = load_workbook(execution.relatorio, data_only=True)["Relatorio"]
    assert sheet.max_row == 3
    assert sheet.cell(3, 7).value == "AGUARDANDO_VALIDACAO_HUMANA"


def test_pipeline_usa_ultimo_resultado_quando_coleta_falha(monkeypatch, tmp_path) -> None:
    settings = _settings(tmp_path)
    breaker = CircuitBreaker(failure_threshold=3)
    monkeypatch.setattr("src.pipeline.collect_all", lambda value: _coleta())
    executar_processo(settings, breaker)
    monkeypatch.setattr(
        "src.pipeline.collect_all",
        lambda value: (_ for _ in ()).throw(RuntimeError("GRP indisponível")),
    )

    execution = executar_processo(settings, breaker)

    assert execution.usou_fallback
    assert len(execution.resultados) == 2
    assert "MRP_FALLBACK_ATIVADO" in settings.alert_file.read_text("utf-8")

