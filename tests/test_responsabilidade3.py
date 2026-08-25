from __future__ import annotations

import json
import logging
from pathlib import Path

import pytest
from openpyxl import load_workbook

from responsabilidade3.adapter import AdapterError, adapt_mrp_records
from responsabilidade3.alerts import InMemoryAlertSink, send_alert
from responsabilidade3.circuit_breaker import CircuitBreaker, CircuitState
from responsabilidade3.fallback import FallbackMRPProvider, StaticMRPProvider
from responsabilidade3.fixtures import (
    FailingMRPProvider,
    mock_invalid_mrp_records,
    mock_mrp_raw_records,
    mock_mrp_results,
)
from responsabilidade3.input_contract import (
    ContractValidationError,
    MRPInputContract,
    MRPInputContractDict,
)
from responsabilidade3.models import MRPResult
from responsabilidade3.report import (
    REPORT_HEADERS,
    ReportGenerationError,
    generate_excel_report,
)
from responsabilidade3.severity import Severity
from responsabilidade3.structured_logging import log_event


PROJECT_ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_PATH = PROJECT_ROOT / "Source" / "modelo_relatorio_necessidades.xlsx"


def test_input_contract_represents_mrp_engine_result() -> None:
    payload: MRPInputContractDict = {
        "material": "MAT001",
        "fornecedor": "Fornecedor A",
        "estoque_atual": 100,
        "necessidade": 50,
        "quantidade_comprar": 50,
    }

    result = MRPInputContract(**payload)

    assert result.material == "MAT001"
    assert result.to_report_row()["Necessidade"] == 50


@pytest.mark.parametrize(
    ("field", "value"),
    [
        ("material", ""),
        ("fornecedor", ""),
        ("estoque_atual", -1),
        ("necessidade", -1),
        ("quantidade_comprar", -1),
        ("prazo_dias", -1),
        ("capacidade", "invalido"),
    ],
)
def test_input_contract_rejects_structurally_invalid_values(
    field: str,
    value: object,
) -> None:
    payload = {
        "material": "MAT001",
        "fornecedor": "Fornecedor A",
        "estoque_atual": 100,
        "necessidade": 50,
        "quantidade_comprar": 50,
        "capacidade": 100,
        "prazo_dias": 5,
    }
    payload[field] = value

    with pytest.raises(ContractValidationError):
        MRPInputContract(**payload)


def test_adapter_accepts_future_mrp_contract_aliases() -> None:
    results = adapt_mrp_records(
        [
            {
                "produto": "MAT010",
                "fornecedor": "Fornecedor X",
                "estoque_atual": "12",
                "necessidade": "34",
                "quantidade_comprar": "55",
            }
        ]
    )

    assert results == [
        MRPResult(
            material="MAT010",
            fornecedor="Fornecedor X",
            estoque_atual=12,
            necessidade=34,
            quantidade_comprar=55,
        )
    ]


def test_adapter_accepts_current_excel_shape_without_calculating_mrp() -> None:
    results = adapt_mrp_records(mock_mrp_raw_records())

    assert results[1].material == "MAT003"
    assert results[1].necessidade == 130
    assert results[1].quantidade_comprar == 130
    assert results[1].observacao == "Formato igual ao relatorio atual"


def test_adapter_rejects_missing_required_fields() -> None:
    with pytest.raises(AdapterError):
        adapt_mrp_records(mock_invalid_mrp_records())


def test_generate_excel_report_uses_existing_template_headers(tmp_path: Path) -> None:
    output_path = tmp_path / "relatorio_necessidades.xlsx"

    generate_excel_report(
        mock_mrp_results(),
        output_path=output_path,
        template_path=TEMPLATE_PATH,
    )

    workbook = load_workbook(output_path, data_only=True)
    worksheet = workbook["Relatorio"]

    assert [cell.value for cell in worksheet[1]] == REPORT_HEADERS
    assert worksheet.cell(row=2, column=1).value == "Fornecedor A"
    assert worksheet.cell(row=2, column=2).value == "Parafuso"
    assert worksheet.cell(row=2, column=4).value == 250
    assert worksheet.max_row == 4


def test_generate_excel_report_with_single_entry_and_none(tmp_path: Path) -> None:
    output_path = tmp_path / "relatorio_uma_linha.xlsx"
    result = MRPInputContract(
        material="MAT010",
        fornecedor="Fornecedor X",
        estoque_atual=0,
        necessidade=1,
        quantidade_comprar=1,
        capacidade=None,
        prazo_dias=None,
    )

    generate_excel_report([result], output_path=output_path, template_path=TEMPLATE_PATH)

    workbook = load_workbook(output_path, data_only=False)
    worksheet = workbook["Relatorio"]

    assert worksheet.max_row == 2
    assert worksheet.cell(row=2, column=5).value is None
    assert worksheet.cell(row=2, column=6).value is None


def test_generate_excel_report_with_empty_list(tmp_path: Path) -> None:
    output_path = tmp_path / "relatorio_vazio.xlsx"

    generate_excel_report([], output_path=output_path, template_path=TEMPLATE_PATH)

    worksheet = load_workbook(output_path)["Relatorio"]
    assert [cell.value for cell in worksheet[1]] == REPORT_HEADERS
    assert worksheet.max_row == 1


def test_generate_excel_report_does_not_overwrite_template(tmp_path: Path) -> None:
    with pytest.raises(ReportGenerationError):
        generate_excel_report(
            mock_mrp_results(),
            output_path=TEMPLATE_PATH,
            template_path=TEMPLATE_PATH,
        )


def test_generate_excel_report_prevents_formula_injection(tmp_path: Path) -> None:
    output_path = tmp_path / "relatorio_formula.xlsx"
    result = MRPInputContract(
        material="=SUM(A1:A10)",
        fornecedor="-Fornecedor",
        estoque_atual=1,
        necessidade=2,
        quantidade_comprar=3,
        observacao="+texto externo",
    )

    generate_excel_report([result], output_path=output_path, template_path=TEMPLATE_PATH)

    worksheet = load_workbook(output_path, data_only=False)["Relatorio"]
    assert worksheet.cell(row=2, column=2).value == "'=SUM(A1:A10)"
    assert worksheet.cell(row=2, column=1).value == "'-Fornecedor"
    assert worksheet.cell(row=2, column=8).value == "'+texto externo"


def test_generate_excel_report_rejects_invalid_entry(tmp_path: Path) -> None:
    output_path = tmp_path / "relatorio_invalido.xlsx"

    with pytest.raises(ReportGenerationError):
        generate_excel_report(
            [{"material": "MAT001"}],
            output_path=output_path,
            template_path=TEMPLATE_PATH,
        )


def test_generate_excel_report_emits_info_log(tmp_path: Path, caplog) -> None:
    output_path = tmp_path / "relatorio_log.xlsx"
    logger = logging.getLogger("tests.report")

    with caplog.at_level(logging.INFO, logger="tests.report"):
        generate_excel_report(
            mock_mrp_results(),
            output_path=output_path,
            template_path=TEMPLATE_PATH,
            logger=logger,
            execution_id="exec-123",
        )

    payload = _json_log_payload(caplog, "report_generated")
    assert payload["level"] == "INFO"
    assert payload["module"] == "excel_report"
    assert payload["execution_id"] == "exec-123"
    assert payload["context"]["row_count"] == 3


def test_fallback_logs_and_publishes_alert_when_primary_fails(caplog) -> None:
    alerts = InMemoryAlertSink()
    provider = FallbackMRPProvider(
        primary=FailingMRPProvider(),
        fallback=StaticMRPProvider(mock_mrp_results()),
        alert_sink=alerts,
        circuit_breaker=CircuitBreaker(failure_threshold=2),
    )

    with caplog.at_level(logging.WARNING):
        results = provider.get_results()

    assert results == mock_mrp_results()
    payload = _json_log_payload(caplog, "mrp_fallback_used")
    assert payload["level"] == "WARNING"
    assert alerts.alerts[0].code == "MRP_FALLBACK_ATIVADO"
    assert alerts.alerts[0].level == "WARNING"


def test_send_alert_redacts_sensitive_context_and_logs(caplog) -> None:
    alerts = InMemoryAlertSink()
    logger = logging.getLogger("tests.alerts")

    with caplog.at_level(logging.WARNING, logger="tests.alerts"):
        alert = send_alert(
            alerts,
            Severity.WARNING,
            "Falha recuperavel",
            context={"password": "nao_logar", "supplier": "Fornecedor A"},
            code="TEST_ALERT",
            logger=logger,
            execution_id="exec-alert",
        )

    assert alert.context["password"] == "[REDACTED]"
    assert alerts.alerts == [alert]
    payload = _json_log_payload(caplog, "alert_sent")
    assert payload["context"]["password"] == "[REDACTED]"
    assert "nao_logar" not in caplog.text


def test_circuit_breaker_opens_and_uses_fallback() -> None:
    alerts = InMemoryAlertSink()
    circuit_breaker = CircuitBreaker(failure_threshold=1, recovery_timeout_seconds=60)
    provider = FallbackMRPProvider(
        primary=FailingMRPProvider(),
        fallback=StaticMRPProvider(mock_mrp_results()),
        alert_sink=alerts,
        circuit_breaker=circuit_breaker,
    )

    provider.get_results()
    provider.get_results()

    assert circuit_breaker.state == CircuitState.OPEN
    assert [alert.code for alert in alerts.alerts] == [
        "MRP_FALLBACK_ATIVADO",
        "MRP_CIRCUIT_OPEN",
    ]


def test_circuit_breaker_open_to_half_open_to_closed(caplog) -> None:
    current_time = 0.0

    def clock() -> float:
        return current_time

    attempts = {"count": 0}
    circuit_breaker = CircuitBreaker(
        failure_threshold=1,
        recovery_timeout=10,
        clock=clock,
        logger=logging.getLogger("tests.circuit.closed"),
    )

    with caplog.at_level(logging.INFO, logger="tests.circuit.closed"):
        with pytest.raises(RuntimeError):
            circuit_breaker.call(lambda: _raise_runtime_error("falha externa"))

        current_time = 11.0
        result = circuit_breaker.call(lambda: attempts.update(count=1) or "ok")

    assert result == "ok"
    assert attempts["count"] == 1
    assert circuit_breaker.state == CircuitState.CLOSED
    assert _json_log_payload(caplog, "circuit_state_changed", "half_open")
    assert _json_log_payload(caplog, "circuit_state_changed", "closed")


def test_circuit_breaker_half_open_failure_returns_to_open(caplog) -> None:
    current_time = 0.0

    def clock() -> float:
        return current_time

    circuit_breaker = CircuitBreaker(
        failure_threshold=1,
        recovery_timeout=10,
        clock=clock,
        logger=logging.getLogger("tests.circuit.open"),
    )

    with caplog.at_level(logging.WARNING, logger="tests.circuit.open"):
        with pytest.raises(RuntimeError):
            circuit_breaker.call(lambda: _raise_runtime_error("primeira falha"))

        current_time = 11.0
        with pytest.raises(RuntimeError):
            circuit_breaker.call(lambda: _raise_runtime_error("segunda falha"))

    assert circuit_breaker.state == CircuitState.OPEN
    assert _json_log_payload(caplog, "circuit_state_changed", "open")


def test_structured_log_includes_expected_fields(caplog) -> None:
    logger = logging.getLogger("tests.structured")

    with caplog.at_level(logging.ERROR, logger="tests.structured"):
        log_event(
            logger,
            level=Severity.ERROR,
            event="step_failed",
            module="tests",
            message="Etapa falhou de forma controlada",
            execution_id="exec-log",
            supplier="Fornecedor A",
            error_type="RuntimeError",
            duration_ms=12,
        )

    payload = _json_log_payload(caplog, "step_failed")
    assert payload["timestamp"]
    assert payload["level"] == "ERROR"
    assert payload["execution_id"] == "exec-log"
    assert payload["supplier"] == "Fornecedor A"
    assert payload["error_type"] == "RuntimeError"
    assert payload["duration_ms"] == 12


def _raise_runtime_error(message: str) -> None:
    raise RuntimeError(message)


def _json_log_payload(caplog, event: str, new_state: str | None = None) -> dict:
    for record in caplog.records:
        try:
            payload = json.loads(record.message)
        except json.JSONDecodeError:
            continue
        if payload.get("event") != event:
            continue
        if new_state is None:
            return payload
        if payload.get("context", {}).get("new_state") == new_state:
            return payload
    raise AssertionError(f"Log estruturado nao encontrado: {event}")
