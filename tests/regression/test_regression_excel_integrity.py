from pathlib import Path
from openpyxl import load_workbook
from src.reporting.excel import generate_excel_report
from src.reporting.models import MRPInputContract


def test_regression_excel_sheet_structure_and_types(tmp_path: Path):
    output_file = tmp_path / "relatorio_regressao.xlsx"
    items = [
        MRPInputContract("MAT001", "Fornecedor A", 100, 100, 100, 200, 5, "COMPRA_SUGERIDA", "Pedido normal"),
        MRPInputContract("MAT002", "Fornecedor B", 300, 0, 0, 400, 7, "SEM_COMPRA", "Estoque OK"),
        MRPInputContract("MAT003", "Fornecedor C", 50, 210, None, 100, 14, "AGUARDANDO_VALIDACAO_HUMANA", "Alfandega"),
        MRPInputContract("MAT004", "Fornecedor D", 120, 10, 10, 250, 6, "COMPRA_SUGERIDA", "Pedido normal"),
    ]

    generate_excel_report(items, output_path=output_file)
    assert output_file.exists()

    wb = load_workbook(output_file, data_only=True)
    assert "Relatorio" in wb.sheetnames
    sheet = wb["Relatorio"]

    # Validar cabeçalhos
    expected_headers = [
        "Fornecedor",
        "Material",
        "Estoque",
        "Necessidade",
        "Capacidade",
        "Prazo_Dias",
        "Status_Validacao",
        "Observacao",
    ]
    actual_headers = [sheet.cell(1, col).value for col in range(1, 9)]
    assert actual_headers == expected_headers

    # Validar total de linhas
    assert sheet.max_row == 5  # 1 header + 4 rows

    # Validar linha 4 (MAT003 - Validação Humana)
    assert sheet.cell(4, 1).value == "Fornecedor C"
    assert sheet.cell(4, 2).value == "MAT003"
    assert sheet.cell(4, 3).value == 50
    assert sheet.cell(4, 4).value == 210
    assert sheet.cell(4, 7).value == "AGUARDANDO_VALIDACAO_HUMANA"
